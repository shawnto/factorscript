import win32api,win32con,win32clipboard
import autopy
import time
import Tkinter as tk
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import takeScreenshot
import locationDatas
#from tkinter import Tk


#Factor Scripts
#Shawn Owen

#global location variables
#"At home" donates standard Mac Pro monitor,
#"At work" values based on current 2nd monitor at desk.
x_cordFileMaint = 865
y_cordFileMaint  = 424
x_cord_AdvCStore = 801
y_cord_AdvCStore  = 456
#x_cord_PriceFileMaint = 620
#y_cord_PriceFileMaint = 602
x_cord_PriceFileMaint = 742
y_cord_PriceFileMaint = 698 # At Work
#x_cord_setFocusClick = 773
#y_cord_setFocusClick = 238 
x_cord_setFocusClick = 974 #At Work
y_cord_setFocusClick = 322
x_cord_CostField = 915
y_cord_CostField = 543
#x_cord_SingleField = 792 #At home
#y_cord_SingleField = 451
x_cord_SingleField = 1035 #At work
y_cord_SingleField = 543
x_cord_MultiField = 1087
y_cord_MultiField = 543
x_cord_CaseField = 1146
y_cord_CaseField = 543
#x_cord_EnteringRetailField = 973 #At home
#y_cord_EnteringRetailField = 451
x_cord_EnteringRetailField = 1201
y_cord_EnteringRetailField = 543
x_PkgGroup = 1018
y_PkgGroup = 431
x_upcNumbers = 1131
y_upcNumbers = 469
#global pricing values
#Price Change Values
#Default values for compiler
vendorNumber = '1111'
zoneNumber = '1111' 
subZone  = '1111'
promoPrice = '1111'
startDate = '1111'
endDate = '1111'  
itemNumber = '111' 
singlePrice = '11'
multiPrice = '111'
casePrice = '111'
enteringRetail = '111' 
cost = '111' 
#Item Add Values
itemDesc = '111'
packingSize = '111'
salesGroup = '111'
salesSubGroup = '0'
packageGroup = '1'
categoryNumber = '402'
upcCase = '0025900227401'
upcUnit = '0025900227401'
unitOfMeasure = 'CS'
foodStamp = 'No'
idRequired = '0'
hasUpc = '0'
vendorNumber = '3696'
itemNumber = '111'
itemDesc = '11'





#x_cord_AddLine = 727
#y_cord_AddLine = 713
#x_cord_EditLine = 787
#y_cord_EditLine = 713

#get data, "type" into field.
def enterData(data):
    data = str(data)
    autopy.key.type_string(data,240)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(.5)
#tap return
def enterOnce():
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(.5)
#Click location, enter data. For sensitive data fields
def enterDataWClick(data,x,y):
    data = str(data)
    autopy.mouse.move(x,y)
    autopy.mouse.click()
    autopy.key.type_string(data,180)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(.5)
    autopy.mouse.click()

#Click @location
def leftClickLocation(x,y):
    autopy.mouse.move(x,y)
    time.sleep(.01)
    autopy.mouse.click()
#as above
def rightClickLocation(x,y):
    autopy.mouse.move(x,y)
    time.sleep(.01)
    autopy.mouse.click(autopy.mouse.RIGHT_BUTTON)
#click, hold, drag release
def dragToLocation(x,y):
    autopy.mouse.toggle(down,mouse.LEFT_BUTTON)
    autopy.mouse.smooth_move(x,y)
    autopy.mouse.toggle(up, mouse.LEFT_BUTTON)
    
#Move over num times. NOTE: USES 'ENTER',
#USE leftClickLocation for sensitive data
def enterMulti(num):
    for i in range(0,num):
        autopy.key.tap(autopy.key.K_RETURN)
        time.sleep(.05)

def validateFieldEditable(location,x,y):
    #Take the image values of the case field.
    temp = takeScreenshot.screenShot(location)
    leftClickLocation(x,y)
    #move the mouse cursor out of the way for next check
    leftClickLocation(1156,454)
    time.sleep(.05)
    temp2 = takeScreenshot.screenShot(location)
    #ensure field is highlighted.
    return (takeScreenshot.compareScreenShot(temp,temp2))

#navigate to file maintanence menu
def toFileMaint():
    leftClickLocation(x_cord_AdvCStore,y_cord_AdvCStore)
    time.sleep(.05)
    leftClickLocation(x_cordFileMaint,y_cordFileMaint)
#navigate to price file maintanence
def toPriceFileMaint():
    leftClickLocation(x_cord_PriceFileMaint,y_cord_PriceFileMaint)
    time.sleep(3.0)
    leftClickLocation(x_cord_setFocusClick,y_cord_setFocusClick)
    time.sleep(.5)
#navigate to item file maintanence
def toItemFileMaint():
    leftClickLocation(x_cord_setFocusClick,y_cord_setFocusClick)
    time.sleep(.5)
    autopy.key.tap('2')
    time.sleep(.1)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(2)

#Grabs the source file, and places them into list.
#then, runs appropriate method for situation.
#All similar methods read, "readSource[some action]"
def readSourceFileAddItem():
    itemInfo = []
   # referenceFile = input("Enter source file")
    refBook = load_workbook('ItemData.xlsx')
    print(refBook.get_sheet_names())
    sheet_ranges = refBook['Sheet 1']
    sheet1 = refBook.active
    index = 1
    for cellObj in sheet1.columns:
        if(index < sheet1.max_row):
            for cellObj in sheet1.rows[index]:
                itemInfo.append(cellObj.value)
                #print(cellObj)
                #print(itemInfo)
                
        index += 1
        #while values in "itemInfo" run process
        if(itemInfo):
            setItemValues(itemInfo)
            addNewItem()
            cont = input("Fix Single Desc")
            leftClickLocation(x_cord_setFocusClick,y_cord_setFocusClick)
            autopy.key.tap('i',autopy.key.MOD_ALT)
            cont = input("Type 1 to continue")
#See above       
def readSourceFileAddPrice():
    toPriceFileMaint()
    itemInfo = []
   # referenceFile = input("Enter source file")
    refBook = load_workbook('ItemData.xlsx')
    print(refBook.get_sheet_names())
    sheet_ranges = refBook['Sheet 1']
    sheet1 = refBook.active
    index = 1
    for rows in sheet1.rows:
        if(index < sheet1.max_row):
            for cellObj in sheet1.rows[index]:
                itemInfo.append(cellObj.value)
                #print(cellObj)
                #print(itemInfo)
        index += 1
        if(itemInfo):
            setItemValues(itemInfo)
            addItemPriceChange()
            #print "Pass: %d" % index
            #print(itemInfo)
            #cont = input("Type 1 to continue")

def readSourceFileEditLine():
    infoToEdit = input("Enter 1 for Cost, 2 for single, 3 for multi, " +
                       "4 for case, and 5 for entering retail.")
    toPriceFileMaint()
    itemInfo = []
   # referenceFile = input("Enter source file")
    refBook = load_workbook('ItemData.xlsx')
    print(refBook.get_sheet_names())
    sheet_ranges = refBook['Sheet 1']
    sheet1 = refBook.active
    index = 1
    for rows in sheet1.rows:
        if(index < sheet1.max_row):
            for cellObj in sheet1.rows[index]:
                itemInfo.append(cellObj.value)
                #print(cellObj)
                #print(itemInfo)
        index += 1
        if(itemInfo):
            setItemValues(itemInfo)
            editLine(infoToEdit)
            #print "Pass: %d" % index
            #print(itemInfo)
            #cont = input("Type 1 to continue")


def readSourceValidateInfo(infoToCheck,upcFlag):
    if(infoToCheck == 6):
        typeOfUpc = input("Enter 1 for Case Upc, 2 for Single...")
        toItemFileMaint()
    itemInfo = []
   # referenceFile = input("Enter source file")
    refBook = load_workbook('ItemData.xlsx')
    print(refBook.get_sheet_names())
    sheet_ranges = refBook['Sheet 1']
    sheet1 = refBook.active
    index = 1
    for rows in sheet1.rows:
        if(index < sheet1.max_row):
            for cellObj in sheet1.rows[index]:
                itemInfo.append(cellObj.value)
                #print(cellObj)
                #print(itemInfo)
        index += 1
        #user selects values to validate, integers donate values to check.
        if(itemInfo):
            setItemValues(itemInfo)
           # print(itemInfo)
            if (infoToCheck == 1):
                testValue = validateCost()
            elif (infoToCheck == 2):
                testValue = validateRetail()
            elif (infoToCheck == 3):
                testValue = validateMulti()
            elif (infoToCheck == 4):
                testValue = validateCase()
            elif (infoToCheck == 5):
                testValue = validateEnteringRetail()
            else:
                testValue = validateUpc(typeOfUpc)
            if(testValue == 1):
                    print itemDesc + " is Correct"
            else:
                    print "ERROR on: " + itemDesc
            #stop = input("STOP")
            autopy.key.tap('c',autopy.key.MOD_ALT)
            time.sleep(1.5)
    autopy.key.tap('x',autopy.key.MOD_ALT)
            
def setItemValues(itemInfoCopy):
    #Price Change Values
    itemInfoCopy.reverse()
    global vendorName
    vendorName = str(itemInfoCopy.pop())
    global vendorNumber
    vendorNumber =  str(itemInfoCopy.pop())
    print(vendorNumber)
    global zoneNumber
    zoneNumber = itemInfoCopy.pop()
    global subZone
    subZone = itemInfoCopy.pop()
    global promoPrice
    promoPrice = itemInfoCopy.pop()
    global startDate
    startDate = itemInfoCopy.pop()
    global endDate
    endDate = itemInfoCopy.pop()
    global itemNumber
    itemNumber = itemInfoCopy.pop()
    global itemDesc
    itemDesc = itemInfoCopy.pop()
    global itemSize
    itemSize = itemInfoCopy.pop()
    global singlePrice
    singlePrice = itemInfoCopy.pop()
    global multiPrice
    multiPrice = itemInfoCopy.pop()
    global casePrice
    casePrice = itemInfoCopy.pop()
    global enteringRetail
    enteringRetail = itemInfoCopy.pop()
    global cost
    cost = itemInfoCopy.pop()
    #Item Add Values
    global hasUpc
    hasUpc = itemInfoCopy.pop()
    global packingSize
    packingSize = itemInfoCopy.pop()
    global salesGroup
    salesGroup = itemInfoCopy.pop()
    global salesSubGroup
    salesSubGroup = itemInfoCopy.pop()
    global packageGroup
    packageGroup = itemInfoCopy.pop()
    global categoryNumber
    categoryNumber = itemInfoCopy.pop()
    global upcCase
    upcCase = itemInfoCopy.pop()
    global upcUnit
    upcUnit = itemInfoCopy.pop()
    global unitOfMeasure
    unitOfMeasure = itemInfoCopy.pop()
    global foodStamp
    foodStamp = itemInfoCopy.pop()
    global idRequired
    idRequired = itemInfoCopy.pop()
    #isCoolerItem is a value for item database project. 
    global isCoolerItem
    isCoolerItem = itemInfoCopy.pop()
    
#for all adds and edits, the appropriate menu must be opened in factor.


#add a "line" for a price change
def addItemPriceChange():
    #Key taps are for menu shortcuts.
    autopy.key.tap('a',autopy.key.MOD_ALT)
    #sleep times are CRUCIAL, factor needs large amounts of recovery time.
    time.sleep(1)
    enterData(vendorNumber)
    time.sleep(.05)
    enterData(zoneNumber)
    time.sleep(.05)
    enterData(subZone)
    time.sleep(.05)
    #important to have correct in source file,
    #the following will "check" or "uncheck" promoprice in factor.
    #will offset command order if incorrect.
    if(promoPrice == '1'):
        autopy.key.tap(autopy.key.K_RIGHT)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(1)
    enterData(startDate)
    if(promoPrice == '1'):
        enterData(endDate)
    enterData(itemNumber)
    if(cost != '0'):
        enterDataWClick(cost,x_cord_CostField,y_cord_CostField)
        time.sleep(.05)
    if(singlePrice != '0'):
        enterDataWClick(singlePrice,x_cord_SingleField,
                        y_cord_SingleField)
        time.sleep(.05)
    if(multiPrice != '0'):
        enterDataWClick(multiPrice,x_cord_MultiField,y_cord_MultiField)
        time.sleep(.05)
    if(casePrice != '0' and validateFieldEditable('Case',x_cord_CaseField,
                                                  y_cord_CaseField)):
        #reset the field, enter the data
        leftClickLocation(x_cord_CostField,y_cord_CostField)
        enterDataWClick(casePrice,x_cord_CaseField,
                        y_cord_CaseField)
        time.sleep(.05)
    if(enteringRetail != '0'):
        #reset the field, enter the data
        leftClickLocation(x_cord_CostField,y_cord_CostField)
        enterDataWClick(enteringRetail,
                        x_cord_EnteringRetailField,
                        y_cord_EnteringRetailField)
        time.sleep(2)
    autopy.key.tap('i',autopy.key.MOD_ALT)
    time.sleep(5)
#edit a line
#reads the appropriate field, and edits the line.
def editLine(infoToChange):
    
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(zoneNumber)
    enterData(subZone)
    time.sleep(.25)
    enterData(itemNumber)
    if(infoToChange == 1):
        if(cost != '0'):
            leftClickLocation(x_cord_SingleField,y_cord_SingleField)
            enterDataWClick(cost,x_cord_CostField,y_cord_CostField)
            time.sleep(.5)
    elif(infoToChange == 2):
        if(singlePrice != '0'):
            enterDataWClick(singlePrice,x_cord_SingleField,
                            y_cord_SingleField)
            time.sleep(.5)
    elif(infoToChange == 3):
        if(multiPrice != '0'):
            enterDataWClick(multiPrice,x_cord_MultiField,y_cord_MultiField)
            time.sleep(.5)
    elif(infoToChange == 4):
        if(casePrice != '0'):
            enterDataWClick(casePrice,x_cord_CaseField,
                            y_cord_CaseField)
            time.sleep(.5)
    else:
        if(enteringRetail != '0'):
            enterDataWClick(enteringRetail,
                            x_cord_EnteringRetailField,
                            y_cord_EnteringRetailField)
            time.sleep(2)
    #pause = input("PAUSE")
    autopy.key.tap('u',autopy.key.MOD_ALT)
    time.sleep(3)
def addNewItem():
    leftClickLocation(x_cord_setFocusClick,y_cord_setFocusClick)
    autopy.key.tap('a',autopy.key.MOD_ALT)
    time.sleep(1.0)
    if(hasUpc == '0'):
        autopy.key.tap(autopy.key.K_RIGHT)
        autopy.key.tap(autopy.key.K_RIGHT)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(.1)
    enterData(vendorNumber)
    enterData(itemNumber)
    enterData(itemDesc)
    autopy.key.tap(autopy.key.K_RETURN)
    autopy.key.tap('n',autopy.key.MOD_ALT)
    enterData(packingSize)
    if(salesGroup == '1'):
        enterData(salesGroup)
    if(salesSubGroup != '0'):
        enterData(salesSubGroup)
    leftClickLocation(x_PkgGroup,y_PkgGroup)
    enterData(packageGroup)
    if (hasUpc == '0'):
        enterMulti(1)
    else:
        enterMulti(2)
    time.sleep(.5)
    enterData(categoryNumber)
    if (hasUpc == '1'):
        leftClickLocation(x_upcNumbers,y_upcNumbers)
        time.sleep(1)
        if(packingSize == '1/1'):
            enterData(upcUnit)
            enterData(upcUnit)
            enterData(itemDesc)
        else:
            enterData(upcCase)
            enterData(upcUnit)
            enterData(itemDesc)
            enterData(upcCase)
            enterData(itemDesc)
        autopy.key.tap('k',autopy.key.MOD_ALT)
        enterMulti(2)
    else:
        enterMulti(1)
    enterData(unitOfMeasure)
    enterMulti(1)
    if(idRequired == '0'):
            enterOnce()
    elif(idRequired == '1'):
            autopy.key.tap(autopy.key.K_DOWN)
            enterOnce()
    else:
            autopy.key.tap(autopy.key.K_DOWN)
            autopy.key.tap(autopy.key.K_DOWN)
            enterOnce()
    if(foodStamp == '1'):
            autopy.key.tap(autopy.key.K_DOWN)
            time.sleep(.5)
    enterMulti(2)

#clik a textfield, and select all in that field.
def toSelectAll(x,y):
    leftClickLocation(x,y)
    time.sleep(.25)
    leftClickLocation(x,y)
    leftClickLocation(x,y)
    rightClickLocation(x,y)
    for i in range(0,6):
        autopy.key.tap(autopy.key.K_DOWN)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(.5)
    autopy.key.tap('c',autopy.key.MOD_CONTROL)
    
#copys to the clipboard.
def pullText(x,y):
    board = tk.Tk()
    toSelectAll(x,y)
    text = str(board.selection_get(selection = "CLIPBOARD"))
    #board.clipboard_clear()
    return text


def toSelectAllItemFile(x,y):
    leftClickLocation(x,y)
    rightClickLocation(x,y)
    for i in range(0,3):
        autopy.key.tap(autopy.key.K_DOWN)
    autopy.key.tap(autopy.key.K_RETURN)
    time.sleep(.5)
    #autopy.key.tap('c',autopy.key.MOD_CONTROL)

def pullTextItemFile(x,y):
    board = tk.Tk()
    toSelectAllItemFile(x,y)
    text = str(board.selection_get(selection = "CLIPBOARD"))
    #board.clipboard_clear()
    return text

#validate[something] will navigate to the item, and verify that info.
#must be specific. 0 != 0.00, for example.
def validateCost():
    toPriceFileMaint()
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(zoneNumber)
    enterData(subZone)
    enterData(itemNumber)
    text = pullText(x_cord_CostField,y_cord_CostField)
    print text
    #print cost
    if (text != cost):
        return 0
    else:
        return 1

def validateRetail():
    toPriceFileMaint()
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(zoneNumber)
    enterData(subZone)
    enterData(itemNumber)
    #toSelectAll()
    text = pullText(x_cord_SingleField,y_cord_SingleField)
    print text
    #print cost
    if (text != singlePrice):
        return 0
    else:
        return 1
def validateEnteringRetail():
    toPriceFileMaint()
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(zoneNumber)
    enterData(subZone)
    enterData(itemNumber)
    text = pullText(x_cord_EnteringRetailField,y_cord_EnteringRetailField)
    print text
    #print cost
    if (text != enteringRetail):
        return 0
    else:
        return 1
def validateMulti():
    toPriceFileMaint()
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(zoneNumber)
    enterData(subZone)
    enterData(itemNumber)
    text = pullText(x_cord_MultiField,y_cord_MultiField)
    #print text
    #print cost
    if (text != multiPrice):
        return 0
    else:
        return 1

def validateCase():
    toPriceFileMaint()
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(zoneNumber)
    enterData(subZone)
    enterData(itemNumber)
    text = pullText(x_cord_CaseField,y_cord_CaseField)
    #print text
    #print cost
    if (text != casePrice):
        return 0
    else:
        return 1

def validateUpc(upcType):
    x_upcCase = 861
    y_upcCase = 625
    x_upcSingle = 863
    y_upcSingle = 560
    autopy.key.tap('e',autopy.key.MOD_ALT)
    time.sleep(.5)
    enterData(vendorNumber)
    enterData(itemNumber)
    autopy.key.tap('n',autopy.key.MOD_ALT)
    time.sleep(1)
    leftClickLocation(x_upcNumbers,y_upcNumbers)
    if (upcType == 1):
        text = pullTextItemFile(x_upcCase,y_upcCase)
    else:
        text = pullTextItemFile(x_upcSingle,y_upcSingle)
    print text
    time.sleep(5)
    autopy.key.tap('c',autopy.key.MOD_ALT)
    #print text
    #print cost
    if (upcType == 1):
        if (text != upcCase):
            return 0
        else:
            return 1
    else:
        if (text != upcUnit):
            return 0
        else:
            return 1
    
#below is helpful to keep in code for pixel hunting
#x,y = autopy.mouse.get_pos()

typeOfProcess = input("Enter 1 for New Item," +
                      "2 for New Price Line, 3 to validate info..., "+
                      "4 to edit...., 5 to exit")
if(typeOfProcess == 1):
    readSourceFileAddItem()
elif(typeOfProcess == 2):
    readSourceFileAddPrice()
elif(typeOfProcess == 3):
    typeOfValidation = input("Enter 1 for Cost, 2 for Single Retail," +
                             " 3 for Multi Retail, 4 for Case Retail, " +
                             "5 for Entering Retail, and 6 for UPC..., " +
                             "7 for the gauntlet, 8 to exit.")
    if(typeOfValidation == 7):
        for i in range(1,7):
            readSourceValidateInfo(i,1)
    elif(typeOfValidation < 7):
        readSourceValidateInfo(typeOfValidation,0)
    else:
        print "CANCELLED"
elif(typeOfProcess == 4):
    readSourceFileEditLine()
else:
    print "CANCELLED"
   

pauseText = input("Enter any Key to exit")
#print x,y <- for pixel hunting.

